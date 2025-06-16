import tkinter as tk
from tkinter import messagebox, simpledialog, ttk, filedialog
import json
import os
import cv2
import requests
import time
import re
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import logging
import threading
from pathlib import Path
from PIL import Image, ImageTk
import io

# Імпорти для генерації штрихкодів
try:
    import barcode
    from barcode.writer import ImageWriter
    BARCODE_AVAILABLE = True
except ImportError:
    BARCODE_AVAILABLE = False
    print("Модуль barcode не встановлено. Встановіть: pip install python-barcode[images]")

# Перевірка доступності OpenCV для RTSP
RTSP_AVAILABLE = True
try:
    import cv2
except ImportError:
    RTSP_AVAILABLE = False
    print("OpenCV не встановлено. Встановіть: pip install opencv-python")

CONFIG_FILE = "config.json"
LOG_FILE = "app.log"

# Отримуємо шлях до робочого столу за замовчуванням
DEFAULT_DESKTOP_PATH = Path.home() / "Desktop"
DEFAULT_SAVE_FOLDER = DEFAULT_DESKTOP_PATH / "SkanerFoto"

# Налаштування логування: в файл і в консоль, з рівнем INFO
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler()
    ]
)

def validate_ip_address(ip):
    """Перевірка коректності IP-адреси"""
    pattern = r'^((25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$'
    return re.match(pattern, ip) is not None

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                config = json.load(f)
                logging.info("Конфігурація завантажена з файлу.")
                return config
        except (json.JSONDecodeError, Exception) as e:
            logging.error(f"Помилка читання конфігурації: {e}")
            return create_default_config()
    else:
        logging.info("Файл конфігурації не знайдено, створена дефолтна структура.")
        return create_default_config()

def create_default_config():
    return {
        "telegram_token": "",
        "telegram_chat_id": "",
        "camera_ip": "",
        "camera_login": "",
        "camera_password": "",
        "recorder_ip": "",
        "recorder_login": "",
        "recorder_password": "",
        "recorder_port": "554",
        "recorder_channel": "1",
        "recorder_rtsp_template": "hikvision",
        "use_recorder": True,  # Використовувати реєстратор замість окремої камери
        "save_folder": str(DEFAULT_SAVE_FOLDER),
        "packers": []
    }

def save_config(config):
    try:
        os.makedirs(os.path.dirname(CONFIG_FILE) if os.path.dirname(CONFIG_FILE) else ".", exist_ok=True)
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=4, ensure_ascii=False)
        logging.info("Конфігурація збережена у файл.")
        return True
    except Exception as e:
        logging.error(f"Помилка збереження конфігурації: {e}")
        return False

def ensure_save_folder(folder_path):
    """Створюємо папку для збереження"""
    try:
        Path(folder_path).mkdir(parents=True, exist_ok=True)
        logging.info(f"Папка для збереження: {folder_path}")
        return True
    except Exception as e:
        logging.error(f"Помилка створення папки збереження: {e}")
        return False

def get_current_save_folder():
    """Отримання поточної папки збереження з конфігурації"""
    config = load_config()
    return config.get("save_folder", str(DEFAULT_SAVE_FOLDER))

def generate_barcode_image(code):
    """Генерація штрихкода в пам'яті (без збереження)"""
    if not BARCODE_AVAILABLE:
        return None
    
    try:
        # Генеруємо штрихкод Code128
        code128 = barcode.get_barcode_class('code128')
        barcode_instance = code128(code, writer=ImageWriter())
        
        # Зберігаємо в BytesIO (в пам'яті)
        buffer = io.BytesIO()
        barcode_instance.write(buffer)
        buffer.seek(0)
        
        # Відкриваємо як PIL Image
        img = Image.open(buffer)
        
        logging.info(f"Штрихкод згенеровано в пам'яті для коду: {code}")
        return img
        
    except Exception as e:
        logging.error(f"Помилка генерації штрихкода: {e}")
        return None

def get_rtsp_templates():
    """Шаблони RTSP URL для різних виробників"""
    return {
        "hikvision": {
            "name": "Hikvision",
            "main": "rtsp://{login}:{password}@{ip}:{port}/Streaming/Channels/{channel}01",
            "sub": "rtsp://{login}:{password}@{ip}:{port}/Streaming/Channels/{channel}02"
        },
        "dahua": {
            "name": "Dahua",
            "main": "rtsp://{login}:{password}@{ip}:{port}/cam/realmonitor?channel={channel}&subtype=0",
            "sub": "rtsp://{login}:{password}@{ip}:{port}/cam/realmonitor?channel={channel}&subtype=1"
        },
        "generic": {
            "name": "Загальний",
            "main": "rtsp://{login}:{password}@{ip}:{port}/live/ch{channel}",
            "sub": "rtsp://{login}:{password}@{ip}:{port}/live/ch{channel}_sub"
        },
        "axis": {
            "name": "Axis",
            "main": "rtsp://{login}:{password}@{ip}:{port}/axis-media/media.amp?videocodec=h264&resolution=1920x1080",
            "sub": "rtsp://{login}:{password}@{ip}:{port}/axis-media/media.amp?videocodec=h264&resolution=640x480"
        },
        "foscam": {
            "name": "Foscam",
            "main": "rtsp://{login}:{password}@{ip}:{port}/videoMain",
            "sub": "rtsp://{login}:{password}@{ip}:{port}/videoSub"
        }
    }

def test_rtsp_connection(ip, port, login, password, template, channel):
    """Тестування RTSP підключення"""
    if not RTSP_AVAILABLE:
        return False, "OpenCV не встановлено. Встановіть: pip install opencv-python"
    
    try:
        templates = get_rtsp_templates()
        
        if template not in templates:
            return False, f"Невідомий шаблон: {template}"
        
        # Генеруємо RTSP URL
        rtsp_data = templates[template]
        rtsp_url = rtsp_data["main"].format(
            login=login, 
            password=password, 
            ip=ip, 
            port=port, 
            channel=channel
        )
        
        logging.info(f"[RTSP Test] Тестування: {rtsp_url}")
        
        # Тестуємо підключення
        cap = cv2.VideoCapture(rtsp_url)
        
        if not cap.isOpened():
            cap.release()
            return False, f"❌ Не вдалося підключитися до RTSP потоку\n\nURL: {rtsp_url}\n\n🔧 Перевірте:\n• IP-адресу та порт\n• Логін та пароль\n• Номер каналу\n• Налаштування мережі"
        
        # Спробуємо прочитати кадр
        ret, frame = cap.read()
        cap.release()
        
        if not ret or frame is None:
            return False, f"❌ RTSP потік підключено, але немає відео\n\nURL: {rtsp_url}\n\n🔧 Можливі причини:\n• Канал вимкнений\n• Проблеми з кодеком\n• Потік зайнятий іншим клієнтом"
        
        height, width = frame.shape[:2]
        info = f"✅ RTSP підключення успішне!\n\n"
        info += f"Виробник: {rtsp_data['name']}\n"
        info += f"URL: {rtsp_url}\n"
        info += f"Розширення: {width}x{height}\n"
        info += f"Канал: {channel}"
        
        return True, info
        
    except Exception as e:
        return False, f"Помилка RTSP тестування: {str(e)}"

def get_rtsp_screenshot(ip, port, login, password, template, channel):
    """Отримання скриншота з RTSP потоку реєстратора"""
    if not RTSP_AVAILABLE:
        logging.error("[RTSP] OpenCV не встановлено")
        return None
    
    try:
        templates = get_rtsp_templates()
        
        if template not in templates:
            logging.error(f"[RTSP] Невідомий шаблон: {template}")
            return None
        
        # Генеруємо RTSP URL (спочатку основний потік, потім субпотік)
        rtsp_data = templates[template]
        rtsp_urls = [
            rtsp_data["main"].format(login=login, password=password, ip=ip, port=port, channel=channel),
            rtsp_data["sub"].format(login=login, password=password, ip=ip, port=port, channel=channel)
        ]
        
        for rtsp_url in rtsp_urls:
            try:
                logging.info(f"[RTSP Screenshot] Спроба: {rtsp_url}")
                
                cap = cv2.VideoCapture(rtsp_url)
                
                if not cap.isOpened():
                    logging.warning(f"[RTSP Screenshot] Не вдалося відкрити: {rtsp_url}")
                    continue
                
                # Читаємо декілька кадрів для стабілізації
                for i in range(5):
                    ret, frame = cap.read()
                    if not ret:
                        break
                
                cap.release()
                
                if ret and frame is not None:
                    # Створити папку temp, якщо не існує
                    save_folder = Path(get_current_save_folder()) / "temp"
                    save_folder.mkdir(exist_ok=True)
                    screenshot_path = save_folder / f"rtsp_screenshot_{int(time.time())}.jpg"
                    
                    # Зберігаємо скриншот
                    cv2.imwrite(str(screenshot_path), frame)
                    
                    if screenshot_path.exists() and screenshot_path.stat().st_size > 1024:
                        logging.info(f"[RTSP Screenshot] Скриншот отримано: {screenshot_path}")
                        return str(screenshot_path)
                    else:
                        if screenshot_path.exists():
                            screenshot_path.unlink()
                        logging.warning(f"[RTSP Screenshot] Файл занадто малий")
                else:
                    logging.warning(f"[RTSP Screenshot] Не вдалося прочитати кадр з {rtsp_url}")
                
            except Exception as e:
                logging.error(f"[RTSP Screenshot] Помилка для {rtsp_url}: {e}")
                continue
        
        logging.error("[RTSP Screenshot] Всі спроби отримання скриншота невдалі")
        return None
        
    except Exception as e:
        logging.error(f"[RTSP Screenshot] Загальна помилка: {e}")
        return None

def send_telegram_message(token, chat_id, message):
    if not token or not chat_id:
        logging.warning("[Telegram] Токен або Chat ID не налаштовані")
        return False
    
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    try:
        response = requests.post(url, data={"chat_id": chat_id, "text": message}, timeout=10)
        if response.status_code == 200:
            logging.info(f"[Telegram] Відправлено повідомлення. Статус: {response.status_code}")
            return True
        else:
            logging.error(f"[Telegram] Помилка відправки. Статус: {response.status_code}")
            return False
    except Exception as e:
        logging.error(f"[Telegram Error] Помилка відправки повідомлення: {e}")
        return False

def send_telegram_photo(token, chat_id, photo_path, caption=""):
    if not token or not chat_id:
        logging.warning("[Telegram] Токен або Chat ID не налаштовані")
        return False
    
    if not os.path.exists(photo_path):
        logging.error(f"[Telegram Photo] Файл не знайдено: {photo_path}")
        return False
    
    url = f"https://api.telegram.org/bot{token}/sendPhoto"
    try:
        with open(photo_path, "rb") as f:
            files = {"photo": f}
            data = {"chat_id": chat_id, "caption": caption}
            response = requests.post(url, files=files, data=data, timeout=30)
            if response.status_code == 200:
                logging.info(f"[Telegram Photo] Фото відправлено. Статус: {response.status_code}")
                return True
            else:
                logging.error(f"[Telegram Photo] Помилка відправки фото. Статус: {response.status_code}")
                return False
    except Exception as e:
        logging.error(f"[Telegram Photo Error] Помилка відправки фото: {e}")
        return False

def test_camera_connection_advanced(ip, login, password):
    """Розширене тестування підключення до камери"""
    if not validate_ip_address(ip):
        return False, "Некоректна IP-адреса"
    
    # Список можливих URL для різних типів камер
    test_urls = [
        f"http://{ip}/cgi-bin/snapshot.cgi",  # Стандартний CGI
        f"http://{ip}/snapshot.cgi",          # Без cgi-bin
        f"http://{ip}/cgi-bin/snapshot.jpg",  # З розширенням
        f"http://{ip}/snapshot.jpg",          # Простий snapshot
        f"http://{ip}/image/jpeg.cgi",        # Axis камери
        f"http://{ip}/jpg/image.jpg",         # Деякі IP камери
        f"http://{ip}/videostream.cgi?rate=0&user={login}&pwd={password}",  # З параметрами в URL
    ]
    
    results = []
    
    for url in test_urls:
        try:
            # Тестуємо різні методи аутентифікації
            auth_methods = [
                ("Basic Auth", requests.auth.HTTPBasicAuth(login, password)),
                ("Digest Auth", requests.auth.HTTPDigestAuth(login, password)),
                ("No Auth", None),
            ]
            
            for auth_name, auth in auth_methods:
                try:
                    response = requests.get(url, auth=auth, timeout=5)
                    status = response.status_code
                    content_type = response.headers.get('content-type', '')
                    
                    result_info = f"{url} [{auth_name}] - Status: {status}"
                    
                    if status == 200:
                        if 'image' in content_type.lower():
                            return True, f"✅ Успіх! {result_info}\nТип контенту: {content_type}"
                        else:
                            result_info += f" (Не зображення: {content_type})"
                    elif status == 401:
                        result_info += " (Помилка аутентифікації)"
                    elif status == 404:
                        result_info += " (URL не знайдено)"
                    elif status == 403:
                        result_info += " (Доступ заборонено)"
                    
                    results.append(result_info)
                    
                except requests.exceptions.Timeout:
                    results.append(f"{url} [{auth_name}] - Таймаут")
                except requests.exceptions.ConnectionError:
                    results.append(f"{url} [{auth_name}] - Помилка підключення")
                except Exception as e:
                    results.append(f"{url} [{auth_name}] - Помилка: {str(e)}")
                    
        except Exception as e:
            results.append(f"{url} - Критична помилка: {str(e)}")
    
    return False, "Не вдалося підключитися до камери.\n\nРезультати тестування:\n" + "\n".join(results)

def get_camera_snapshot_advanced(ip, login, password):
    """Покращена функція отримання знімків з підтримкою різних методів"""
    if not validate_ip_address(ip):
        logging.error(f"[Snapshot] Некоректна IP-адреса: {ip}")
        return None
    
    # Спробуємо різні URL та методи аутентифікації
    urls_to_try = [
        f"http://{ip}/cgi-bin/snapshot.cgi",
        f"http://{ip}/snapshot.cgi",
        f"http://{ip}/cgi-bin/snapshot.jpg",
        f"http://{ip}/snapshot.jpg",
        f"http://{ip}/image/jpeg.cgi",
    ]
    
    auth_methods = [
        ("Basic", requests.auth.HTTPBasicAuth(login, password)),
        ("Digest", requests.auth.HTTPDigestAuth(login, password)),
    ]
    
    for url in urls_to_try:
        for auth_name, auth in auth_methods:
            try:
                logging.info(f"[Snapshot] Спроба: {url} з {auth_name} Auth")
                response = requests.get(url, auth=auth, stream=True, timeout=10)
                
                if response.status_code == 200:
                    content_type = response.headers.get('content-type', '')
                    if 'image' in content_type.lower():
                        # Створити папку temp, якщо не існує
                        save_folder = Path(get_current_save_folder())
                        temp_folder = save_folder / "temp"
                        temp_folder.mkdir(parents=True, exist_ok=True)
                        img_path = temp_folder / f"snap_{int(time.time())}.jpg"
                        
                        with open(img_path, "wb") as f:
                            for chunk in response.iter_content(1024):
                                f.write(chunk)
                        
                        logging.info(f"[Snapshot] Знімок отримано: {img_path} (метод: {auth_name}, URL: {url})")
                        return str(img_path)
                    else:
                        logging.warning(f"[Snapshot] Отримано не зображення: {content_type}")
                else:
                    logging.warning(f"[Snapshot] HTTP {response.status_code} для {url} з {auth_name}")
                    
            except requests.exceptions.Timeout:
                logging.warning(f"[Snapshot] Таймаут для {url} з {auth_name}")
            except requests.exceptions.ConnectionError:
                logging.warning(f"[Snapshot] Помилка підключення для {url} з {auth_name}")
            except Exception as e:
                logging.error(f"[Snapshot] Помилка {url} з {auth_name}: {e}")
    
    logging.error("[Snapshot] Всі спроби отримання знімка невдалі")
    return None

def cleanup_temp_files():
    """Очищення старих тимчасових файлів"""
    try:
        save_folder = Path(get_current_save_folder())
        temp_folder = save_folder / "temp"
        if temp_folder.exists():
            for file_path in temp_folder.iterdir():
                if file_path.is_file():
                    # Видалити файли старші 1 години
                    if time.time() - file_path.stat().st_ctime > 3600:
                        file_path.unlink()
                        logging.info(f"Видалено старий тимчасовий файл: {file_path}")
    except Exception as e:
        logging.error(f"Помилка очищення тимчасових файлів: {e}")

def cleanup_old_files():
    """Очищення файлів старше 2 тижнів"""
    try:
        save_folder = Path(get_current_save_folder())
        two_weeks_ago = datetime.now() - timedelta(weeks=2)
        deleted_count = 0
        
        # Очищення папки sessions
        sessions_folder = save_folder / "sessions"
        if sessions_folder.exists():
            for session_path in sessions_folder.iterdir():
                if session_path.is_dir():
                    # Перевіряємо дату створення папки
                    if datetime.fromtimestamp(session_path.stat().st_ctime) < two_weeks_ago:
                        # Видаляємо всі файли в папці
                        for file_path in session_path.iterdir():
                            if file_path.is_file():
                                file_path.unlink()
                                deleted_count += 1
                        # Видаляємо саму папку, якщо вона порожня
                        try:
                            session_path.rmdir()
                            logging.info(f"Видалено папку сесії: {session_path}")
                        except OSError:
                            logging.warning(f"Папка не видалена (не порожня): {session_path}")
                    else:
                        # Перевіряємо файли в папці
                        for file_path in session_path.iterdir():
                            if file_path.is_file():
                                if datetime.fromtimestamp(file_path.stat().st_ctime) < two_weeks_ago:
                                    file_path.unlink()
                                    deleted_count += 1
        
        # Очищення папки temp
        temp_folder = save_folder / "temp"
        if temp_folder.exists():
            for file_path in temp_folder.iterdir():
                if file_path.is_file():
                    if datetime.fromtimestamp(file_path.stat().st_ctime) < two_weeks_ago:
                        file_path.unlink()
                        deleted_count += 1
        
        logging.info(f"Очищення завершено. Видалено файлів: {deleted_count}")
        return deleted_count
        
    except Exception as e:
        logging.error(f"Помилка очищення старих файлів: {e}")
        return -1

def log_to_excel(folder, barcode, timestamp):
    try:
        # Створити папку, якщо не існує
        folder_path = Path(folder)
        folder_path.mkdir(parents=True, exist_ok=True)
        filename = folder_path / "session_log.xlsx"
        
        if filename.exists():
            wb = load_workbook(filename)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Час", "Штрихкод"])
        
        ws.append([timestamp, barcode])
        wb.save(filename)
        logging.info(f"[Excel] Запис до журналу: {barcode} о {timestamp}")
        return True
    except Exception as e:
        logging.error(f"[Excel] Помилка при записі в Excel: {e}")
        return False

class BarcodeProcessor:
    def __init__(self, config):
        self.config = config
        self.current_packer = None
        self.session_folder = None
        logging.info("Ініціалізовано BarcodeProcessor.")

    def process_code(self, code):
        code = code.strip()
        if not code:
            logging.warning("Отримано порожній код")
            return "Порожній код"
        
        now_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        logging.info(f"Отримано код: {code}")

        # Перевірка на ID пакувальника (3 цифри)
        if len(code) == 3 and code.isdigit():
            packer = next((p for p in self.config["packers"] if p["id"] == code), None)
            if packer:
                self.current_packer = packer
                # Створити папку sessions
                save_folder = Path(self.config.get("save_folder", str(DEFAULT_SAVE_FOLDER)))
                sessions_folder = save_folder / "sessions"
                sessions_folder.mkdir(parents=True, exist_ok=True)
                self.session_folder = sessions_folder / f"{now_str}_{packer['name']}_{packer['id']}"
                try:
                    self.session_folder.mkdir(exist_ok=True)
                    logging.info(f"Обрано пакувальника: {packer['name']} (ID {packer['id']}). Створена сесійна папка {self.session_folder}")
                    send_telegram_message(self.config["telegram_token"], self.config["telegram_chat_id"],
                                          f"🧑‍🏭 Пакувальник {packer['name']} (#{packer['id']}) почав роботу.")
                    return f"Обрано пакувальника: {packer['name']}"
                except Exception as e:
                    logging.error(f"Помилка створення папки сесії: {e}")
                    return "Помилка створення папки сесії"
            else:
                logging.warning(f"Пакувальник з ID {code} не знайдений.")
                return f"Пакувальник з ID {code} не знайдений"

        # Обробка штрихкода товару
        if self.current_packer:
            return self.process_product_barcode(code)
        else:
            logging.warning("Спочатку проскануйте ID пакувальника!")
            return "Спочатку проскануйте ID пакувальника!"

    def process_product_barcode(self, code):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Вибираємо джерело зображення: реєстратор або окрема камера
        snapshot_path = None
        
        if self.config.get("use_recorder", True) and self.config.get("recorder_ip"):
            # Отримуємо скриншот з реєстратора через RTSP
            logging.info("[Snapshot] Використовуємо реєстратор для скриншотів")
            snapshot_path = get_rtsp_screenshot(
                self.config["recorder_ip"],
                self.config.get("recorder_port", "554"),
                self.config["recorder_login"],
                self.config["recorder_password"],
                self.config.get("recorder_rtsp_template", "hikvision"),
                self.config.get("recorder_channel", "1")
            )
        elif self.config.get("camera_ip"):
            # Отримуємо скриншот з окремої камери
            logging.info("[Snapshot] Використовуємо окрему камеру для скриншотів")
            snapshot_path = get_camera_snapshot_advanced(
                self.config["camera_ip"],
                self.config["camera_login"],
                self.config["camera_password"]
            )
        
        if snapshot_path:
            filename = f"{code}_{timestamp.replace(':', '-').replace(' ', '_')}.jpg"
            final_path = self.session_folder / filename
            
            try:
                img = cv2.imread(snapshot_path)
                if img is not None:
                    # Додати текст на зображення
                    cv2.putText(img, f"{code} {timestamp}", (10, 30), cv2.FONT_HERSHEY_SIMPLEX,
                                1, (0, 0, 255), 2, cv2.LINE_AA)
                    
                    # Додати інформацію про джерело
                    source_info = "Реєстратор" if self.config.get("use_recorder", True) else "Камера"
                    cv2.putText(img, f"Джерело: {source_info}", (10, 70), cv2.FONT_HERSHEY_SIMPLEX,
                                0.7, (0, 255, 0), 2, cv2.LINE_AA)
                    
                    cv2.imwrite(str(final_path), img)
                    
                    # Видалити тимчасовий файл
                    try:
                        Path(snapshot_path).unlink()
                    except:
                        pass
                    
                    logging
